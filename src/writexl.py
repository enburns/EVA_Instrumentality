#!/bin/usr/python2
import openpyxl
import project

def write_proj_tab(proj, tab, proj_count):
	tab.cell(row=proj_count+1, column=1, value=proj.title)
	tab.cell(row=proj_count+1, column=2, value=proj.alias)
	tab.cell(row=proj_count+1, column=3, value=proj.description)
	tab.cell(row=proj_count+1, column=4, value=proj.center)
	tab.cell(row=proj_count+1, column=5, value=proj.taxid)

def write_analy_tab(analy, tab, analy_count):
	tab.cell(row=analy_count+1, column=1, value=analy.title)
	tab.cell(row=analy_count+1, column=2, value=analy.alias)
	tab.cell(row=analy_count+1, column=3, value=analy.description)
	tab.cell(row=analy_count+1, column=4, value=analy.prj.title)
	tab.cell(row=analy_count+1, column=5, value=analy.experiment)
	tab.cell(row=analy_count+1, column=6, value=analy.ref)
	tab.cell(row=analy_count+1, column=7, value=analy.refmd5)

def write_file_tab(file, tab, file_count):
	tab.cell(row=file_count+1, column=1, value=file.analysis.alias)
	tab.cell(row=file_count+1, column=2, value=file.name)
	tab.cell(row=file_count+1, column=3, value=file.filetype)
	tab.cell(row=file_count+1, column=4, value=file.md5)

def write(read_user, user_infos, projects, out):
	wb = openpyxl.load_workbook(filename='src/EVA_Submission_template.V1.1.0.xlsx')
	if read_user:
		ws = wb['Submitter Details']
		ws['A2'] = user_infos['LAST_NAME']
		ws['B2'] = user_infos['FIRST_NAME']
		ws['C2'] = user_infos['TELEPHONE']
		ws['D2'] = user_infos['EMAIL']
		ws['E2'] = user_infos['LABORATORY']
		ws['F2'] = user_infos['CENTER']
		ws['G2'] = user_infos['ADDRESS']
	project_tab = wb['Project']
	analysis_tab = wb['Analysis']
	file_tab = wb['Files']
	proj_count = 1
	analy_count = 1
	file_count = 1
	for proj in projects:
		write_proj_tab(proj, project_tab, proj_count)
		proj_count += 1
		for analy in proj.analyses:
			write_analy_tab(analy, analysis_tab, analy_count)
			analy_count += 1
			for f in analy.files:
				write_file_tab(f, file_tab, file_count)
				file_count += 1
	wb.save(out)
